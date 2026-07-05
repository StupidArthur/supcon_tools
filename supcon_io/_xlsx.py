"""
xlsx 读写(openpyxl)。

默认约定:
- 固定读/写第一个 sheet
- 数字 cell 处理默认 forbid(命中数字 cell 抛 ExcelPrecisionError)
- 数字/浮点/日期时间等原生类型透出
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Literal

from .types import Table

NumericHandling = Literal["forbid", "allow", "force_text"]


class ExcelPrecisionError(ValueError):
    """Excel 数字 cell 处理策略设为 'forbid' 时,遇到数字 cell 抛出。"""


def _normalize_excel_handling(value: str) -> NumericHandling:
    if value not in ("forbid", "allow", "force_text"):
        raise ValueError(
            f"excel_numeric_handling 必须是 forbid/allow/force_text,got {value!r}"
        )
    return value  # type: ignore[return-value]


# ───────── read ─────────


def _read_xlsx(  # noqa: PLR0913
    path: str | Path,
    *,
    header_rows: int | None = None,
    excel_numeric_handling: str = "forbid",
    sniff: bool = True,  # xlsx 暂不用 sniff 做什么,只是对称签名保留
) -> Table:
    handling = _normalize_excel_handling(excel_numeric_handling)
    eff_header_rows = header_rows if header_rows is not None else 1

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "读 xlsx 需要 openpyxl: pip install openpyxl"
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]  # 固定第一个 sheet
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return Table(title=[], desc=None, data=[])

    # header_rows=0:无表头
    if eff_header_rows == 0:
        # 全部当数据
        if handling == "forbid":
            for row in rows:
                _check_no_numbers(row, handling)
        return Table(title=[], desc=None, data=[list(r) for r in rows])

    # 第 1 行表头
    title_row = rows[0]
    title = [(v if v is not None else "") for v in title_row]

    desc: list[str] | None = None
    data_start = 1
    if eff_header_rows >= 2 and len(rows) >= 2:
        desc_row = rows[1]
        desc = [(v if v is not None else "") for v in desc_row]
        data_start = 2

    data = [list(r) for r in rows[data_start:]]

    if handling == "forbid":
        for row in data:
            _check_no_numbers(row, handling)

    return Table(title=list(title), desc=desc, data=data)


def _check_no_numbers(row, handling: NumericHandling) -> None:
    """forbid 模式下,数据行里不允许出现非 str/None 的值。"""
    if handling != "forbid":
        return
    for v in row:
        # 允许:None、str、bool(bool 也是 int 的子类,需先判)
        if v is None or isinstance(v, str):
            continue
        if isinstance(v, bool):
            # bool 是 int 但语义上是"开/关",保留
            continue
        if isinstance(v, (int, float)):
            raise ExcelPrecisionError(
                f"Excel 数字 cell 命中 forbid 策略,值={v!r}。"
                "如需容忍 Excel 浮点精度损失,显式传 excel_numeric_handling='allow' "
                "或 'force_text'。"
            )


# ───────── write ─────────


def _write_xlsx(  # noqa: PLR0913
    path: str | Path,
    table: Table,
    *,
    sheet_name: str = "Sheet1",
    header_rows: int = 1,
) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError(
            "写 xlsx 需要 openpyxl: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    eff_header_rows = header_rows
    if table.desc is None:
        eff_header_rows = 1

    if table.title:
        ws.append(table.title)
    if eff_header_rows >= 2 and table.desc is not None:
        ws.append(table.desc)
    for row in table.data:
        ws.append([("" if v is None else v) for v in row])

    wb.save(filename=str(p))
