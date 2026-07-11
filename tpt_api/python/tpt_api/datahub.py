"""ibd-data-hub-web-v2.2 tag + 历史值管理（与 data-hub-tool/common_api.py 1:1 对齐）。

端点（POST/DELETE 混合，统一 base URL）：
- /ibd-data-hub-web-v2.2/api/tag-info/add                     注册位号
- /ibd-data-hub-web-v2.2/api/tag-info/page                    分页列位号 (MyBatis Page)
- /ibd-data-hub-web-v2.2/api/tag-info/batchDeleteLogic        批量逻辑删除
- /ibd-data-hub-web-v2.2/api/tag-info/batchDelete             批量物理删除
- /ibd-data-hub-web-v2.2/api/tag-info/batchUpdate             批量修改位号参数（节点/单位/采集周期）
- /ibd-data-hub-web-v2.2/api/tag-info/update                  编辑单个位号（全字段可改）
- /ibd-data-hub-web-v2.2/api/tag-info/export                 导出位号（返回 Excel 文件 blob）
- /ibd-data-hub-web-v2.2/api/tag-info/importTagInfoStream    导入位号（上传 Excel 文件，multipart）
- /ibd-data-hub-web-v2.2/api/tag-info/getNotUsedBaseTagInfoContinue  查数据源未导入的位号（游标分页）
- /ibd-data-hub-web-v2.2/api/tag-info/batchAdd               从数据源批量导入位号
- /ibd-data-hub-web-v2.2/api/tag-group/queryWithQuality      查位号（带实时值+质量码，支持多条件过滤）
- /ibd-data-hub-web-v2.2/api/tag-group/get                    回收站按 groupId 查
- /ibd-data-hub-web-v2.2/api/tag-group/add                   创建位号分组节点
- /ibd-data-hub-web-v2.2/api/tag-group/update                编辑位号分组节点
- /ibd-data-hub-web-v2.2/api/tag-group/batchDelete           删除位号分组节点（isForce 控制是否同时删位号）
- /ibd-data-hub-web-v2.2/api/tag-group/groupTree             获取位号分组节点树
- /ibd-data-hub-web-v2.2/api/tag-group/batchAddRelation      收藏位号（关联位号到分组）
- /ibd-data-hub-web-v2.2/api/tag-group/batchDelRelation      取消收藏（移除位号与分组的关联）
- /ibd-data-hub-web-v2.2/api/tag-value/importTagValue         同步 JSON 历史值
- /ibd-data-hub-web-v2.2/api/tag-value/importTagValueHistory  异步 Excel/ZIP 历史值
- /ibd-data-hub-web-v2.2/api/tag-value/importCSVTagValueHistory  CSV（已废弃）
- /ibd-data-hub-web-v2.2/api/tag-value/getHistoryValueFromDB  历史值查询（验证）
- /ibd-data-hub-web-v2.2/api/tag-value/collectTagValue        位号值采集（触发采集任务）
- /ibd-data-hub-web-v2.2/api/tag-value/getRTValue             取位号实时值（List<数据模型>）
- /ibd-data-hub-web-v2.2/api/tag-value/getHistoryValue        取位号历史值（IPage 分页，支持采样/偏移）
- /ibd-data-hub-web-v2.2/api/tag-value/writeTagValues         实时数据库回写位号值
- /ibd-data-hub-web-v2.2/api/ds-info/test                    数据源测试（枚举/读/写/历史，testType 区分）
"""

from __future__ import annotations

import logging
import os
from typing import Any, Callable

from .client import AlgAPI
from .errors import SuccessCode
from .types import DataTypes, DefaultTagTypesAll, DsInfo, DsSubTypes, DsTypes, TagTypes

log = logging.getLogger(__name__)


# data-hub 端点常量。
DataHubBasePath = "/ibd-data-hub-web-v2.2/api"
DataHubTagAdd = DataHubBasePath + "/tag-info/add"
DataHubTagPage = DataHubBasePath + "/tag-info/page"
DataHubTagBatchDeleteLogic = DataHubBasePath + "/tag-info/batchDeleteLogic"
DataHubTagBatchDelete = DataHubBasePath + "/tag-info/batchDelete"
DataHubTagBatchUpdate = DataHubBasePath + "/tag-info/batchUpdate"
DataHubTagUpdate = DataHubBasePath + "/tag-info/update"
DataHubTagExport = DataHubBasePath + "/tag-info/export"
DataHubTagImportStream = DataHubBasePath + "/tag-info/importTagInfoStream"
DataHubTagGetNotUsed = DataHubBasePath + "/tag-info/getNotUsedBaseTagInfoContinue"
DataHubTagBatchAdd = DataHubBasePath + "/tag-info/batchAdd"
DataHubTagGroupQueryWithQuality = DataHubBasePath + "/tag-group/queryWithQuality"
DataHubTagGroupGet = DataHubBasePath + "/tag-group/get"
DataHubTagGroupAdd = DataHubBasePath + "/tag-group/add"
DataHubTagGroupUpdate = DataHubBasePath + "/tag-group/update"
DataHubTagGroupBatchDelete = DataHubBasePath + "/tag-group/batchDelete"
DataHubTagGroupTree = DataHubBasePath + "/tag-group/groupTree"
DataHubTagGroupBatchAddRelation = DataHubBasePath + "/tag-group/batchAddRelation"
DataHubTagGroupBatchDelRelation = DataHubBasePath + "/tag-group/batchDelRelation"
DataHubImportTagValue = DataHubBasePath + "/tag-value/importTagValue"
DataHubImportTagValueHistory = DataHubBasePath + "/tag-value/importTagValueHistory"
DataHubImportCSVTagValueHistory = DataHubBasePath + "/tag-value/importCSVTagValueHistory"
DataHubGetHistoryValueFromDB = DataHubBasePath + "/tag-value/getHistoryValueFromDB"
DataHubCollectTagValue = DataHubBasePath + "/tag-value/collectTagValue"
DataHubGetRTValue = DataHubBasePath + "/tag-value/getRTValue"
DataHubGetHistoryValue = DataHubBasePath + "/tag-value/getHistoryValue"
DataHubWriteTagValues = DataHubBasePath + "/tag-value/writeTagValues"
# 数据源（ds-info）
DataHubDsInfoPage = DataHubBasePath + "/ds-info/page"
DataHubDsInfoAdd = DataHubBasePath + "/ds-info/add"
DataHubDsInfoChangeState = DataHubBasePath + "/ds-info/changeState"
DataHubDsInfoBatchDelete = DataHubBasePath + "/ds-info/batchDelete"
DataHubDsInfoTest = DataHubBasePath + "/ds-info/test"


# === 位号管理 ===

def add_tag(
    api: AlgAPI,
    tag_name: str,
    data_type: int = DataTypes["DOUBLE"],
    tag_type: int = TagTypes["一次位号"],
    ds_id: int = 2,
    group_id: str = "0",
    unit: str = "",
    only_read: bool = False,
    frequency: int = 10,
    need_push: bool = True,
    tag_desc: str | None = None,
    is_vector: bool = True,
    tag_base_name: str | None = None,
    hi_eu: float | None = None,
    lo_eu: float | None = None,
    limit_up: float | None = None,
    limit_up_up: float | None = None,
    limit_up_up_up: float | None = None,
    limit_down: float | None = None,
    limit_down_down: float | None = None,
    limit_down_down_down: float | None = None,
) -> dict[str, Any]:
    """注册一个位号（目前仅支持一次位号；二次位号、虚位号待补充）。

    参数:
      tag_name:       系统位号名（用户面看到的名字），必填
      data_type:      数据类型代码 (1=BOOLEAN .. 13=DATE_TIME)，默认 11=DOUBLE
      tag_type:       位号类型 (1=一次位号, 4=虚位号)，默认 1
      ds_id:          数据源 ID（默认 2 = "我的数据源"）
      group_id:       位号分组 ID，默认 "0" = Root
      unit:           单位
      only_read:      是否只读
      frequency:      采集频率（秒）
      need_push:      是否需要推送
      tag_desc:       描述，默认 "{tag_name} 描述"
      is_vector:      是否向量
      tag_base_name:  底层位号名（指向 OPC UA / 数据源上的实际节点）。
                      默认 = tag_name。
                      绑定到 OPC UA 数据源时，约定格式是 "{namespace_index}_{node_name}"，
                      例如 ns=1 + node="loop_demo_1" -> "1_loop_demo_1"。
                      这时 tagName 可以是 "1_loop_demo_1"（同名）或用户友好的别名。
      hi_eu:          量程上限（可选）
      lo_eu:          量程下限（可选）
      limit_up:           高限（可选）
      limit_up_up:        高高限（可选）
      limit_up_up_up:     高高高限（可选）
      limit_down:         低限（可选）
      limit_down_down:    低低限（可选）
      limit_down_down_down: 低低低限（可选）

    返回: _request 响应 dict（通常含 success/code/msg，content 字段含新建位号信息）。
    """
    if tag_desc is None:
        tag_desc = f"{tag_name} 描述"
    base_name = tag_base_name if tag_base_name is not None else tag_name
    data: dict[str, Any] = {
        "tagType": tag_type,
        "dsId": ds_id,
        "tagBaseName": base_name,
        "tagName": tag_name,
        "dataType": data_type,
        "unit": unit,
        "onlyRead": only_read,
        "frequency": frequency,
        "needPush": need_push,
        "tagDesc": tag_desc,
        "isVector": is_vector,
        "groupId": group_id,
    }
    if hi_eu is not None:
        data["hiEU"] = hi_eu
    if lo_eu is not None:
        data["loEU"] = lo_eu
    if limit_up is not None:
        data["limitUp"] = limit_up
    if limit_up_up is not None:
        data["limitUpUp"] = limit_up_up
    if limit_up_up_up is not None:
        data["limitUpUpUp"] = limit_up_up_up
    if limit_down is not None:
        data["limitDown"] = limit_down
    if limit_down_down is not None:
        data["limitDownDown"] = limit_down_down
    if limit_down_down_down is not None:
        data["limitDownDownDown"] = limit_down_down_down
    body = {"data": data}
    return api._request("POST", DataHubTagAdd, body=body, wrap=False)


def list_tags(
    api: AlgAPI,
    page: int = 1,
    page_size: int = 10,
    sort: str = "-createTime",
    data: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """分页列位号（MyBatis Page 结构，POST /ibd-data-hub-web-v2.2/api/tag-info/page）。

    参数:
      page:      页码（从 1 开始）
      page_size: 每页条数
      sort:      排序字段（如 "-createTime"）
      data:      过滤条件 dict（如 {"dsId": 36, "tagType": 1}）

    返回: content dict（MyBatis Page 结构），含 records / total / size / current / pages / orders。
          每条 record 含 id / tagName / tagBaseName / dataType / dsId 等字段（不带实时值）。
    """
    body: dict[str, Any] = {
        "data": data or {},
        "requestBase": {
            "page": f"{page}-{page_size}",
            "sort": sort,
        },
    }
    return api._request("POST", DataHubTagPage, body=body, wrap=False)


def get_all_tags(
    api: AlgAPI,
    page_size: int = 200,
    sort: str = "-createTime",
    data: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """自动翻页拉取所有位号，缓存到 api.tags 和 api.name_map。

    参数:
      page_size: 每页条数
      sort:      排序字段
      data:      过滤条件 dict

    返回: list[dict]，所有位号记录。
          副作用: api.tags / api.name_map 被缓存。
    """
    all_records: list[dict[str, Any]] = []
    page = 1
    while True:
        result = list_tags(api, page=page, page_size=page_size, sort=sort, data=data)
        records = result.get("records", [])
        if not records:
            break
        all_records.extend(records)
        if len(records) < page_size:
            break
        page += 1
    api.tags = all_records
    api.name_map = {t.get("tagName"): t for t in all_records if t.get("tagName")}
    log.info("get_all_tags 完成: 共 %d tags (filter=%s)", len(all_records), data)
    return all_records


def get_tag_by_name(api: AlgAPI, tag_name: str) -> dict[str, Any] | None:
    """通过 tagName 获取缓存的位号信息（依赖 get_all_tags 先跑过）。

    参数:
      tag_name: 系统位号名

    返回: 位号 dict，未找到返回 None。
    """
    return api.name_map.get(tag_name)


def get_all_tags_all_types(
    api: AlgAPI,
    page_size: int = 2000,
    tag_types: tuple[int, ...] | list[int] = DefaultTagTypesAll,
) -> list[dict[str, Any]]:
    """拉取全部位号，遍历所有 tagType 合并去重（按 id）。

    参数:
      page_size: 每页条数
      tag_types: 要遍历的 tagType 集合，默认 DefaultTagTypesAll

    返回: list[dict]，去重后的位号记录。
          副作用: api.tags / api.name_map 被缓存。

    注意: get_all_tags() 默认 data 为空时平台只返回默认类，会漏掉其它 tagType 的位号。
    本方法逐个 tagType 拉取后合并，确保不漏。
    """
    seen: set[Any] = set()
    all_records: list[dict[str, Any]] = []
    for tt in tag_types:
        try:
            records = get_all_tags(api, page_size=page_size, data={"tagType": tt})
        except Exception:
            records = []
        for t in records:
            tid = t.get("id")
            if tid is not None and tid not in seen:
                seen.add(tid)
                all_records.append(t)
    api.tags = all_records
    api.name_map = {t.get("tagName"): t for t in all_records if t.get("tagName")}
    return all_records


def delete_tags(api: AlgAPI, ids: int | list[int]) -> dict[str, Any]:
    """批量逻辑删除位号（软删，进回收站）（DELETE /api/tag-info/batchDeleteLogic）。

    参数:
      ids: int / int 的可迭代对象

    返回: _parse_resp dict，含 status_code / code / msg / is_success / data / raw。
          注意：逻辑删除，数据保留但位号在正常列表中不再可见，进回收站。
    """
    if isinstance(ids, int):
        ids = [ids]
    ids = [int(i) for i in ids]
    r = api.client.request("DELETE", DataHubTagBatchDeleteLogic, json={"data": {"ids": ids}})
    return api._parse_resp(r)


def delete_tags_physical(api: AlgAPI, ids: int | list[int]) -> dict[str, Any]:
    """物理删除位号（清回收站）（DELETE /api/tag-info/batchDelete）。

    参数:
      ids: int 或 int 的可迭代对象

    返回: _parse_resp dict。

    与 delete_tags（batchDeleteLogic 软删）的区别：这个是物理删，不可恢复。
    """
    if isinstance(ids, int):
        ids = [ids]
    ids = [int(i) for i in ids]
    r = api.client.request("DELETE", DataHubTagBatchDelete, json={"data": {"ids": ids}})
    return api._parse_resp(r)


def delete_tags_by_name(
    api: AlgAPI,
    tag_names: str | list[str],
    refresh: bool = False,
) -> dict[str, Any]:
    """按位号名批量删除（软删）。内部用 name_map 查 id 再调 delete_tags。

    参数:
      tag_names: 单个名字或名字列表
      refresh:   True=先调 get_all_tags 刷新缓存

    返回: {"deleted": list[str], "missing": list[str], "result": resp or None}
    """
    if isinstance(tag_names, str):
        tag_names = [tag_names]
    if refresh or not api.name_map:
        get_all_tags(api)
    deleted: list[str] = []
    missing: list[str] = []
    ids: list[int] = []
    for name in tag_names:
        t = api.name_map.get(name)
        if t and t.get("id") is not None:
            ids.append(int(t["id"]))
            deleted.append(name)
        else:
            missing.append(name)
    result = None
    if ids:
        result = delete_tags(api, ids)
    return {"deleted": deleted, "missing": missing, "result": result}


def batch_update_tags(
    api: AlgAPI,
    tag_ids: list[int],
    group_id: str | None = None,
    unit: str | None = None,
    frequency: int | None = None,
    tag_type: int | None = None,
) -> dict[str, Any]:
    """批量修改位号参数（POST /ibd-data-hub-web-v2.2/api/tag-info/batchUpdate）。

    只能改：分组节点(groupId)、单位(unit)、采集周期(frequency)。
    其他参数（tagName/dataType/dsId 等）不可通过此接口修改。

    参数:
      tag_ids:   位号 ID 列表
      group_id:  目标分组 ID（移动到指定节点）
      unit:      单位
      frequency: 采集周期（秒）
      tag_type:  位号类型（可选）

    返回: list[dict]，每个元素为更新后的完整位号记录，含 id/tagName/tagBaseName/
          unit/frequency/dataType/hiEU/loEU/limitUp/limitDown/onlyRead 等全字段
    """
    data: dict[str, Any] = {"tagIds": tag_ids}
    if tag_type is not None:
        data["tagType"] = tag_type
    if group_id is not None:
        data["groupId"] = group_id
    if unit is not None:
        data["unit"] = unit
    if frequency is not None:
        data["frequency"] = frequency
    body = {"data": data}
    return api._request("POST", DataHubTagBatchUpdate, body=body, wrap=False)


def update_tag(
    api: AlgAPI,
    tag_id: int,
    tag_name: str,
    data_type: int,
    tag_type: int = 1,
    ds_id: int = 2,
    group_id: str = "0",
    unit: str = "",
    only_read: bool = False,
    frequency: int | str = 10,
    need_push: bool = True,
    tag_desc: str | None = None,
    is_vector: bool = True,
    tag_base_name: str | None = None,
    hi_eu: float | None = None,
    lo_eu: float | None = None,
    limit_up: float | None = None,
    limit_up_up: float | None = None,
    limit_up_up_up: float | None = None,
    limit_down: float | None = None,
    limit_down_down: float | None = None,
    limit_down_down_down: float | None = None,
) -> dict[str, Any]:
    """编辑单个位号（PUT /ibd-data-hub-web-v2.2/api/tag-info/update）。

    全量更新：tagName/dataType 必填（与 add_tag 一致），仅 tag_id 标识要改哪条。
    注意：未传的可选字段会被重置为默认值（如 tag_base_name 不传则变回 = tag_name），
    因此编辑时需传入所有需要保留的字段，不能只传修改项。
    与 batch_update_tags 区别：此接口可改所有字段（含 tagName/dataType/dsId/tagBaseName 等）。

    参数:
      tag_id:       位号 ID（必填）
      tag_name:     系统位号名（必填）
      data_type:    数据类型（必填）
      其余参数同 add_tag

    返回: 更新后的完整位号记录，含 id/tagName/tagBaseName/tagDesc/tagType/dsId/unit/
          dataType/baseDataType/onlyRead/frequency/isVector/needPush/updateTime/updateBy
    """
    if tag_desc is None:
        tag_desc = f"{tag_name} 描述"
    base_name = tag_base_name if tag_base_name is not None else tag_name
    data: dict[str, Any] = {
        "id": tag_id,
        "tagName": tag_name,
        "dataType": data_type,
        "tagType": tag_type,
        "dsId": ds_id,
        "groupId": group_id,
        "unit": unit,
        "onlyRead": only_read,
        "frequency": frequency,
        "needPush": need_push,
        "tagDesc": tag_desc,
        "isVector": is_vector,
        "tagBaseName": base_name,
    }
    if hi_eu is not None:
        data["hiEU"] = hi_eu
    if lo_eu is not None:
        data["loEU"] = lo_eu
    if limit_up is not None:
        data["limitUp"] = limit_up
    if limit_up_up is not None:
        data["limitUpUp"] = limit_up_up
    if limit_up_up_up is not None:
        data["limitUpUpUp"] = limit_up_up_up
    if limit_down is not None:
        data["limitDown"] = limit_down
    if limit_down_down is not None:
        data["limitDownDown"] = limit_down_down
    if limit_down_down_down is not None:
        data["limitDownDownDown"] = limit_down_down_down
    body = {"data": data}
    return api._request("PUT", DataHubTagUpdate, body=body, wrap=False)


def export_tags(
    api: AlgAPI,
    tag_ids: list[int],
    interval: int = 0,
    save_path: str | None = None,
    parse: bool = True,
) -> list[list] | bytes | str:
    """导出位号（POST /ibd-data-hub-web-v2.2/api/tag-info/export）。

    下载 Excel 文件并可选解析为 List[List]。导出的本质是为导入做准备。

    Excel 列结构（21 列）:
      Tag Name / Base Tag Name / Tag Type / Datasource Name / Unit / Data Type /
      Expression / Tag Value / Frequency / High Limit / HH Limit / HHH Limit /
      Low Limit / LL Limit / LLL Limit / Description / Group Name /
      Real-time Push / Readonly / Lo EU / Hi EU

    参数:
      tag_ids:    要导出的位号 ID 列表
      interval:   时间间隔（0=导出当前配置）
      save_path:  保存路径（可选，设了就存文件）
      parse:      True=解析 Excel 返回 List[List]（首行为表头）；False=返回 raw bytes

    返回:
      parse=True  -> List[List]，每个子 list 是一行，首行为表头
      parse=False -> bytes（raw 文件内容）
      save_path 设了也会存文件，但返回值仍按 parse 决定

    注意: 位号需有采集数据才能导出成功，否则服务端返回 500 错误。
    """
    import io

    body = {
        "data": {"id_in": ",".join(str(i) for i in tag_ids), "interval": interval},
        "requestBase": {"page": "0-0", "sort": "-createTime"},
    }
    content = api._download("POST", DataHubTagExport, body=body, wrap=False)
    if save_path:
        with open(save_path, "wb") as f:
            f.write(content)
    if not parse:
        return content
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def import_tags_from_file(
    api: AlgAPI,
    file_path: str,
    conflict_strategy: int = 1,
) -> dict[str, Any]:
    """从 Excel 文件导入位号（POST /ibd-data-hub-web-v2.2/api/tag-info/importTagInfoStream）。

    上传 Excel 文件（与 export_tags 导出的格式一致），批量导入/更新位号。

    参数:
      file_path:         Excel 文件路径
      conflict_strategy: 冲突策略（0=跳过已存在的，1=覆盖已存在的）

    返回: {success: bool, code: str}（code="00000" 表示成功）
    """
    with open(file_path, "rb") as f:
        file_bytes = f.read()

    files = {"file": (file_path.rsplit("/", 1)[-1].rsplit("\\", 1)[-1], file_bytes,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {"conflictStrategy": str(conflict_strategy)}

    url = f"{api.base_url}/{DataHubTagImportStream.lstrip('/')}"
    r = api.client.request("POST", url, files=files, data=data)
    r.raise_for_status()
    result = r.json()
    if result.get("code") != SuccessCode:
        from .errors import TptAPIError
        raise TptAPIError(result.get("code", ""), result.get("msg", ""))
    return result.get("content", result)


def get_not_used_tags(
    api: AlgAPI,
    ds_id: int,
    tag_name: str = "",
    continue_id: str = "",
    page: int = 1,
    page_size: int = 1000,
    sort: str = "tagName",
) -> dict[str, Any]:
    """查数据源中未导入的位号（POST /ibd-data-hub-web-v2.2/api/tag-info/getNotUsedBaseTagInfoContinue）。

    从 UA server browse 节点，预期排除已注册的位号，返回未导入的节点列表。
    支持游标分页（continueID），用于大量节点时连续查询。

    参数:
      ds_id:       数据源 ID
      tag_name:    位号名过滤（模糊匹配，空=不过滤）
      continue_id: 游标 ID（首页为空，后续从上次返回的 continueID 继续）
      page:        页码
      page_size:   每页条数
      sort:        排序字段

    返回: dict，含:
      - successTagNames: list[str]
      - successes: list[dict]，每项含 name/description/tagDataType/tagDataTypeName/
                   dataSourceId/readOnly/hubDataType
      - total: int（未导入位号总数）
      - continueID: str（游标，用于续查下一页）
      - pageNum/pageSize: 分页信息
      - isAllSuccess: bool

    注意: 实测导入后立即查询，已导入的位号仍可能出现在列表中（可能有缓存延迟）。
    """
    body = {
        "data": {"dsId": ds_id, "tagName": tag_name, "continueID": continue_id},
        "requestBase": {"page": f"{page}-{page_size}", "sort": sort},
    }
    return api._request("POST", DataHubTagGetNotUsed, body=body, wrap=False)


def batch_add_tags(
    api: AlgAPI,
    tag_infos: list[dict[str, Any]],
    conflict_strategy: int = 1,
) -> dict[str, Any]:
    """从数据源批量导入位号（POST /ibd-data-hub-web-v2.2/api/tag-info/batchAdd）。

    配合 get_not_used_tags 使用：先查未导入位号，选中后批量导入。

    参数:
      tag_infos:         位号信息列表，每项含:
                         groupId/dsId/tagDesc/dataType/tagType/baseDataType/
                         tagBaseName/tagName/frequency/isVector
      conflict_strategy: 冲突策略（0=跳过，1=覆盖）

    返回: list[dict]，每个元素为创建的位号记录，含 id/tagName/tagBaseName/tagDesc/
          tagType/dsId/dataType/baseDataType/frequency/isVector/needPush/createTime/updateTime
    """
    body = {"data": {"tagInfos": tag_infos, "conflictStrategy": conflict_strategy}}
    return api._request("POST", DataHubTagBatchAdd, body=body, wrap=False)


def query_tags_with_quality(
    api: AlgAPI,
    ds_id: int | None = None,
    group_id: str = "0",
    tag_name: str = "",
    tag_base_name: str = "",
    tag_type: int = 1,
    page: int = 1,
    page_size: int = 100,
    sort: str = "-createTime",
) -> dict[str, Any]:
    """查位号（带实时值+质量码）（POST /ibd-data-hub-web-v2.2/api/tag-group/queryWithQuality）。

    与 list_tags（tag-info/page）区别：返回字段含 tagValue/quality/tagTime/dsName/
    groupName/isCollect 等实时信息，且支持 tagName/tagBaseName 模糊过滤。

    参数:
      ds_id:        数据源 ID（可选）
      group_id:     分组 ID（可选，"0"=Root）
      tag_name:     系统位号名模糊匹配（空=不过滤）
      tag_base_name: 底层位号名模糊匹配（空=不过滤）
      tag_type:     位号类型（1=一次位号）
      page:         页码
      page_size:    每页条数
      sort:         排序字段

    返回: 分组对象，含 id/groupName/tagInfoList{records[],total}。
          每个 record 含 id/tagName/tagBaseName/tagDesc/tagType/dsId/dsName/dataType/
          dataTypeName/tagValue/tagTime/quality/frequency/isCollect/isVector/needPush/groupName
    """
    data: dict[str, Any] = {
        "tagName": tag_name,
        "tagBaseName": tag_base_name,
        "tagType": tag_type,
        "sortField": sort,
        "sortType": 1,
    }
    if ds_id is not None:
        data["dsId"] = ds_id
    data["groupId"] = group_id
    body = {
        "data": data,
        "requestBase": {"page": f"{page}-{page_size}", "sort": sort},
    }
    return api._request("POST", DataHubTagGroupQueryWithQuality, body=body, wrap=False)


# === 回收站 ===

def list_recycle_tags(
    api: AlgAPI,
    page: int = 1,
    page_size: int = 100,
    group_id: str = "1",
    tag_type: int = TagTypes["一次位号"],
    sort: str = "-createTime",
) -> dict[str, Any]:
    """查指定分组下的位号（POST /api/tag-group/get），单页。
    平台回收站用 group_id="1"，收藏夹用 group_id="2"，Root 用 "0"。

    参数:
      page:      页码（从 1 开始）
      page_size: 每页条数
      group_id:  分组 ID（"0"=Root, "1"=回收站, "2"=收藏）
      tag_type:  位号类型（默认 一次位号）
      sort:      排序字段

    返回: 分组对象 dict，含 id / groupName / tagInfoList{records[], total, size, current, orders, ...}。
          每条 record 字段随场景不同（回收站/收藏/Root 内容略有差异）。
    """
    body: dict[str, Any] = {
        "data": {
            "groupId": str(group_id),
            "tagType": tag_type,
            "sortField": sort,
            "sortType": 1,
        },
        "requestBase": {
            "page": f"{page}-{page_size}",
            "sort": sort,
        },
    }
    return api._request("POST", DataHubTagGroupGet, body=body, wrap=False)


def list_favorite_tags(
    api: AlgAPI,
    page: int = 1,
    page_size: int = 100,
    tag_type: int = TagTypes["一次位号"],
    sort: str = "-createTime",
) -> dict[str, Any]:
    """查收藏位号列表（POST /api/tag-group/get, groupId="2"）。

    参数:
      page:      页码
      page_size: 每页条数
      tag_type:  位号类型
      sort:      排序字段

    返回: 分组对象，含 id="2" / groupName="Favorites" / tagInfoList{records[], total, ...}。
          每条 record 含 isCollect=true，以及 tagName / dataType / dataTypeName / dsName /
          tagValue / quality / unit / frequency / hiEU / loEU / limitUp / limitDown 等全字段。
    """
    return list_recycle_tags(api, page=page, page_size=page_size,
                             group_id="2", tag_type=tag_type, sort=sort)


def get_all_recycle_tags(
    api: AlgAPI,
    page_size: int = 100,
    group_id: str = "1",
    tag_type: int = TagTypes["一次位号"],
    on_page: Callable[[int, int], None] | None = None,
) -> list[dict[str, Any]]:
    """翻页拉取回收站/收藏全部位号，返回 list[dict]。

    参数:
      page_size: 每页条数
      group_id:  分组 ID（"1"=回收站, "2"=收藏）
      tag_type:  位号类型
      on_page:   可选回调 on_page(page, accumulated)，每拉一页调用

    返回: list[dict]，所有位号记录。

    注意：tag-group/get 的响应结构是 content.tagInfoList.records
    （位号藏在分组对象的 tagInfoList 里，不是顶层 records）。
    """
    all_records: list[dict[str, Any]] = []
    page = 1
    while True:
        result = list_recycle_tags(api, page=page, page_size=page_size,
                                   group_id=group_id, tag_type=tag_type)
        info = result.get("tagInfoList", {}) if isinstance(result, dict) else {}
        records = info.get("records", [])
        if not records:
            break
        all_records.extend(records)
        if on_page:
            on_page(page, len(all_records))
        if len(records) < page_size:
            break
        page += 1
    return all_records


# === 历史值导入 ===
#
# 注意：3 个导入端点都返回 HTTP 200 / code=00000 ≠ 数据落地，
# 实际写入是异步的，必须用 list_tags / get_history_value 回头查。

def import_tag_value(
    api: AlgAPI,
    data: list[dict[str, Any]],
    ds_id: int | None = None,
) -> dict[str, Any]:
    """JSON 批量导入历史值（同步，一次最多 10000 条，POST /api/tag-value/importTagValue）。

    参数:
      data:  list[dict]，每个 dict 至少含 tagName/tagValue，可选 quality/tagTime/appTime
            时间格式 yyyy-MM-dd HH:mm:ss（空格分隔）
      ds_id: 数据源 ID，None=默认时序库

    返回: _parse_resp dict，含 status_code / code / msg / is_success / data / raw。
    """
    url = DataHubImportTagValue
    body: dict[str, Any] = {"data": data}
    if ds_id is not None:
        body["dsId"] = ds_id
    r = api.client.post(url, json=body)
    return api._parse_resp(r)


def import_tag_value_history(
    api: AlgAPI,
    file_path: str,
    ds_id: int | None = None,
    start_time: str | None = None,
    end_time: str | None = None,
    frequency: int | None = None,
    cron: str | None = None,
) -> dict[str, Any]:
    """Excel / ZIP 导入历史值（异步，POST /api/tag-value/importTagValueHistory）。

    参数:
      file_path: Excel(.xlsx/.xls) 或 ZIP 文件路径
      ds_id:     数据源 ID，None 或 0=默认时序库
      start_time: 起始时间 yyyy-MM-dd HH:mm:ss
      end_time:   结束时间
      frequency:  采样频率（秒）
      cron:       调度表达式；API form 字段名是 corn（拼写错误）而非 cron

    返回: _parse_resp dict。

    注意:
      - HTTP 200 / code=00000 ≠ 数据落地，实际写入是异步的
      - 验证导入结果必须用 list_tags / get_history_value 回头查
      - Excel A1 四段逗号：A1: startTime,endTime,frequency,corn；A2: (空)；
        A3 起: 时间, 位号1值, 位号2值, ...
    """
    ext = os.path.splitext(file_path)[1].lower()
    mime_map = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".zip": "application/zip",
    }
    mime = mime_map.get(ext, "application/octet-stream")

    url = DataHubImportTagValueHistory
    size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
    log.info("上传文件: %s (size=%d bytes, mime=%s)", file_path, size, mime)
    with open(file_path, "rb") as f:
        files = {"file": (os.path.basename(file_path), f, mime)}
        form: dict[str, str] = {}
        if ds_id is not None and ds_id != 0:
            form["dsId"] = str(ds_id)
        if start_time:
            form["startTime"] = start_time
        if end_time:
            form["endTime"] = end_time
        if frequency is not None:
            form["frequency"] = str(frequency)
        if cron:
            form["corn"] = cron  # API 拼写是 corn 不是 cron
        r = api.client.post(url, files=files, data=form)
    result = api._parse_resp(r)
    log.info(
        "上传响应: status=%s code=%s msg=%s requestId=%s",
        result["status_code"], result["code"], result["msg"],
        (result["raw"] or {}).get("requestId") if result.get("raw") else None,
    )
    return result


def import_csv_tag_value_history(api: AlgAPI, file_path: str) -> dict[str, Any]:
    """CSV 导入历史值（POST /api/tag-value/importCSVTagValueHistory）。

    参数:
      file_path: CSV 文件路径

    返回: _parse_resp dict。

    注意: 该接口已废弃，但平台仍可用。新场景请用 import_tag_value_history(Excel/ZIP)。
    """
    url = DataHubImportCSVTagValueHistory
    with open(file_path, "rb") as f:
        files = {"file": (os.path.basename(file_path), f, "text/csv")}
        r = api.client.post(url, files=files)
    return api._parse_resp(r)


# === 历史值查询（用于 import 后的"验证闭环"） ===

def get_history_value(
    api: AlgAPI,
    tag_names: list[str],
    beg_time: str,
    end_time: str,
    is_source: bool = True,
    number_to_string: bool = False,
    page: int = 1,
    page_size: int = 100,
    sort: str = "-appTime",
) -> dict[str, Any]:
    """查位号历史值（单页，POST /api/tag-value/getHistoryValueFromDB）。

    参数:
      tag_names:         系统位号名 list
      beg_time:          起始时间 yyyy-MM-dd HH:mm:ss
      end_time:          结束时间
      is_source:         是否包含源数据（含回写记录），默认 True
      number_to_string:  数值是否转为字符串，默认 False
      page:              页码
      page_size:         每页条数
      sort:              排序字段

    返回: dict，结构 {tagName: {"list": [data points], "total": int, "pageNum": int, "pageSize": int, "totalPage": int}}。
          每个 data point 字段：tagName/tagValue/tagTime/appTime/quality/createTime。
    """
    body: dict[str, Any] = {
        "data": {
            "tagNames": list(tag_names),
            "begTime": beg_time,
            "endTime": end_time,
            "isSource": is_source,
            "numberToString": number_to_string,
        },
        "requestBase": {
            "page": f"{page}-{page_size}",
            "sort": sort,
        },
    }
    return api._request("POST", DataHubGetHistoryValueFromDB, body=body, wrap=False)


def get_all_history(
    api: AlgAPI,
    tag_names: list[str],
    beg_time: str,
    end_time: str,
    is_source: bool = True,
    number_to_string: bool = False,
    page_size: int = 2000,
) -> dict[str, list[dict[str, Any]]]:
    """翻页拉取所有历史值，返回 {tagName: [data points]}。

    参数:
      tag_names:        系统位号名 list
      beg_time:         起始时间
      end_time:         结束时间
      is_source:        是否包含源数据
      number_to_string: 数值是否转为字符串
      page_size:        每页条数

    返回: dict[tagName -> list[data points]]，每个 tag 的所有数据点（按 API 返回顺序，固定最新在前），自动按 total 翻页。
    """
    result: dict[str, list[dict[str, Any]]] = {name: [] for name in tag_names}
    page = 1
    while True:
        page_data = get_history_value(
            api, tag_names=tag_names, beg_time=beg_time, end_time=end_time,
            is_source=is_source, number_to_string=number_to_string,
            page=page, page_size=page_size, sort="-appTime",
        )
        any_remaining = False
        for tag_name, info in page_data.items():
            lst = info.get("list", [])
            result[tag_name].extend(lst)
            if len(result[tag_name]) < info.get("total", 0):
                any_remaining = True
        if not any_remaining:
            break
        page += 1
    return result


# === 位号值采集 / 实时值 / 历史值(分页) / 回写 ===
#
# 以下 4 个端点对应 swagger operationId：
#   collectTagValueUsingPOST  位号值采集（触发采集任务）
#   getRTValueUsingPOST       取位号实时值（List<数据模型>）
#   getHistoryValueUsingPOST  取位号历史值（IPage<数据模型>，支持采样/偏移）
#   writeTagValuesUsingPOST   实时数据库回写位号值
# 4 个端点同属 tag-value 资源，路径前缀与 getHistoryValueFromDB 一致。
# 「数据模型」点字段：tagName/tagValue/tagTime/appTime/quality/dataType/dsId/
#                     id/isSuccess/message/cacheNum/createTime

def collect_tag_value(
    api: AlgAPI,
    es_dto: dict[str, Any],
    group_id: int,
    tenant_id: str,
) -> bool:
    """位号值采集（POST /api/tag-value/collectTagValue）。

    触发/配置一次位号值采集任务。esDTO 是采集任务调度对象（EsDTO），字段较多，
    这里以 dict 透传，由调用方按需组装。

    参数:
      es_dto:    采集任务调度对象（EsDTO）dict，常见字段：
                 taskId / jobName / jobType / cronExpression / executeWay /
                 scheduleType / fixRate / intervalSeconds / startTime /
                 startTimeMillis / tagGroup(TagGroupDTO) / tenantId / xxlJobId /
                 batchId / executeStatus / executeTime / logType / content / msg
      group_id:  位号组 id
      tenant_id: 节点 id

    返回: bool（响应 content），True=采集任务已受理。
    """
    body: dict[str, Any] = {
        "esDTO": es_dto,
        "groupId": group_id,
        "tenantId": tenant_id,
    }
    return api._request("POST", DataHubCollectTagValue, body=body, wrap=True)


def get_rt_value(
    api: AlgAPI,
    tag_names: list[str] | None = None,
    tag_info_ids: list[int] | None = None,
    group_id: int | None = None,
    is_from_db: bool = False,
    option: int | None = None,
    query_time: str | None = None,
) -> list[dict[str, Any]]:
    """取位号实时值（POST /api/tag-value/getRTValue）。

    支持通过位号组 id、位号 id、系统位号名查询，三字段都非必填，但全部为空
    或查不到位号信息会报错。

    参数:
      tag_names:    系统位号名 list（与 tag_info_ids / group_id 三选一或多选）
      tag_info_ids: 位号 id list
      group_id:     位号组 id
      is_from_db:   是否从数据库读取实时值（默认 False，走实时库）
      option:       采样策略，必须与 query_time 配合；query_time 为空时忽略
      query_time:   查询时间 yyyy-MM-dd HH:mm:ss；不传=查最新值。
                    部分数据源不支持指定时间，会在结果中标记失败，不影响其它位号

    返回: list[dict]，每条是一个位号实时值点（数据模型），字段：
          tagName / tagValue / tagTime / appTime / quality / dataType /
          dsId / id / isSuccess / message / cacheNum / createTime
    """
    data: dict[str, Any] = {"isFromDB": is_from_db}
    if tag_names is not None:
        data["tagNames"] = list(tag_names)
    if tag_info_ids is not None:
        data["tagInfoIds"] = list(tag_info_ids)
    if group_id is not None:
        data["groupId"] = group_id
    if option is not None:
        data["option"] = option
    if query_time is not None:
        data["queryTime"] = query_time
    return api._request("POST", DataHubGetRTValue, body=data, wrap=True)


def query_history_value(
    api: AlgAPI,
    tag_names: list[str],
    beg_time: str,
    end_time: str,
    interval: int = 0,
    is_second: bool = True,
    is_source: bool = False,
    offset: int = 0,
    option: int = 0,
    page: int = 1,
    page_size: int = 10,
    sort: str = "-appTime",
) -> dict[str, Any]:
    """取位号历史值（POST /api/tag-value/getHistoryValue，IPage 分页）。

    与 get_history_value（getHistoryValueFromDB）的区别：
      - 本接口支持 interval 采样间隔 / offset 时间偏移 / option 采样填充规则
      - 响应是 MyBatis IPage 结构（records / total / current / size / pages），
        而 get_history_value 返回 {tagName: {list, total}} 结构
      - 起始与结束时间间隔不能超过一个月

    参数:
      tag_names:   系统位号名称 list
      beg_time:    起始时间 yyyy-MM-dd HH:mm:ss
      end_time:    结束时间 yyyy-MM-dd HH:mm:ss
      interval:    采样间隔（秒），0=不采样。设为 N 后每 N 秒抽一条，控制返回条数
      is_second:   是否从第二数据源（时序库）查，默认 True。
                   False=按位号所对应数据源调历史接口（仅一次位号+有数据源有效）
      is_source:   是否查询源数据（含回写记录），默认 False
      offset:      时间偏移量（秒），默认 0。用于调整前值/插值采样的时间范围
      option:      采样填充规则，默认 0=前向采样；isSecond=true 时 3=不填充
      page:        页码（从 1 开始）
      page_size:   每页条数
      sort:        排序字段（如 "-appTime"）

    返回: IPage dict（records / total / current / size / pages），
          records 每条是位号值点（同 get_rt_value 的数据模型字段）。
    """
    body: dict[str, Any] = {
        "data": {
            "tagNames": list(tag_names),
            "begTime": beg_time,
            "endTime": end_time,
            "interval": interval,
            "isSecond": is_second,
            "isSource": is_source,
            "offset": offset,
            "option": option,
        },
        "requestBase": {
            "page": f"{page}-{page_size}",
            "sort": sort,
        },
    }
    return api._request("POST", DataHubGetHistoryValue, body=body, wrap=False)


def write_tag_values(
    api: AlgAPI,
    values: dict[str, Any],
    tag_time: str | None = None,
    quality_code: int | None = None,
) -> dict[str, Any]:
    """实时数据库回写位号值（POST /api/tag-value/writeTagValues）。

    将一组位号值回写到实时数据库，values 是 {tagName: tagValue} 的 map，
    所有位号共用同一个 tagTime / qualityCode。

    参数:
      values:       {tagName: tagValue} dict，一次回写多个位号
      tag_time:     位号时间 yyyy-MM-dd HH:mm:ss，不传=服务端当前时间
      quality_code: 质量码

    返回: WriteTagValuesResp dict：
          - tagNames: list[str]           成功回写的位号名
          - failMsg:  {tagName: errmsg}   失败位号及原因
          - msg:      str                 总体信息
          failMsg 为空且 tagNames 覆盖全部输入=全部成功。
    """
    data: dict[str, Any] = {"values": values}
    if tag_time is not None:
        data["tagTime"] = tag_time
    if quality_code is not None:
        data["qualityCode"] = quality_code
    return api._request("POST", DataHubWriteTagValues, body=data, wrap=True)


# === 数据源 (ds-info) ===
#
# ds-info 描述 ibd-data-hub 连接的下游数据源（OPC-UA-Server / Real time database 等）。
# 注意：ds-id 是 tag 的 dsId 字段指向的目标。

def list_ds_info(
    api: AlgAPI,
    page: int = 1,
    page_size: int = 10,
    sort: str = "-createTime",
    data: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """分页列数据源（MyBatis Page 结构）。

    响应 content 字段是 MyBatis Page（records / total / size / current / pages / orders）。
    每条 record 含 id / name / dsName / dsType / dsTypeDesc / dsSubType / dsSubTypeDesc /
    dsTarUrl / dsStatus / alive / supportSub / dsExtInfo 等。

    参数:
      page:       页码（从 1 开始）
      page_size:  每页条数
      sort:       排序字段（如 "-createTime"）
      data:       过滤条件 dict（实测默认空 data 即可列全）

    返回: content dict（MyBatis Page 结构），含 records / total / size / current / pages / orders。
          每条 record 含 id / name / dsName / dsType / dsSubType / dsTarUrl / dsStatus /
          alive / supportSub / dsExtInfo 等字段。
    """
    if page < 1:
        page = 1
    if page_size < 1:
        page_size = 10
    body: dict[str, Any] = {
        "data": data or {},
        "requestBase": {
            "page": f"{page}-{page_size}",
            "sort": sort,
        },
    }
    return api._request("POST", DataHubDsInfoPage, body=body, wrap=False)


def get_all_ds_info(
    api: AlgAPI,
    page_size: int = 200,
    sort: str = "-createTime",
    data: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """自动翻页拉取所有数据源，返回 list[dict]。

    参数:
      page_size: 每页条数
      sort:      排序字段
      data:      过滤条件 dict

    返回: list[dict]，所有数据源记录。
    """
    if page_size < 1:
        page_size = 200
    all_records: list[dict[str, Any]] = []
    page = 1
    while True:
        result = list_ds_info(api, page=page, page_size=page_size, sort=sort, data=data)
        records = result.get("records", [])
        if not records:
            break
        all_records.extend(records)
        if len(records) < page_size:
            break
        page += 1
    log.info("get_all_ds_info 完成: 共 %d (filter=%s)", len(all_records), data)
    return all_records


def get_ds_info_by_id(
    all_records: list[dict[str, Any]],
    ds_id: int,
) -> dict[str, Any] | None:
    """从 list[dict] 里按 id 查单条数据源（依赖 get_all_ds_info 先跑过）。

    参数:
      all_records: get_all_ds_info 返回的列表
      ds_id:       数据源 ID

    返回: 数据源 dict，未找到返回 None。
    """
    for r in all_records:
        if r.get("id") == ds_id:
            return r
    return None


def get_ds_info_by_name(
    all_records: list[dict[str, Any]],
    ds_name: str,
) -> dict[str, Any] | None:
    """从 list[dict] 里按 dsName 或 name 查单条数据源。

    参数:
      all_records: get_all_ds_info 返回的列表
      ds_name:     数据源名（同时匹配 dsName 和 name 字段）

    返回: 数据源 dict，未找到返回 None。
    """
    for r in all_records:
        if r.get("dsName") == ds_name or r.get("name") == ds_name:
            return r
    return None


def ds_info_to_model(d: dict[str, Any]) -> DsInfo:
    """dict → DsInfo dataclass 便捷方法。

    参数:
      d: 数据源 dict

    返回: DsInfo 实例。
    """
    return DsInfo.from_dict(d)


def add_ds_info(
    api: AlgAPI,
    ds_name: str,
    ds_type: int = DsTypes["REAL_TIME_DB"],
    ds_sub_type: int = DsSubTypes["OPC_UA_SERVER"],
    ds_tar_url: str = "",
    name: str | None = None,
    ds_ext_info: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """新增一个接入数据源（POST /ibd-data-hub-web-v2.2/api/ds-info/add）。

    参数:
      ds_name:     数据源名（必填，全局唯一性由平台按 dsTarUrl 判定）
      ds_type:     数据源大类 (1=Real time database)，默认 1
      ds_sub_type: 数据源子类 (4=OPC-UA-Server)，默认 4
      ds_tar_url:  目标 URL（如 opc.tcp://host:port），必填且全局唯一
      name:        显示名，默认与 ds_name 一致
      ds_ext_info: 扩展信息 dict，默认 {}

    注意:
      - 平台接受 dsType / dsSubType 为字符串（"1" / "4"），本方法内部自动转
      - 必填项缺一不可
      - dsTarUrl 重复 → A0001 「Client error:Duplicate data source address」
      - 响应 content 是新建的完整记录（含 id / createTime / updateTime / createBy / updateBy）

    返回: 新建记录的 content dict（含 id）
    """
    if not ds_name:
        raise ValueError("ds_name 必填")
    if not ds_tar_url:
        raise ValueError("ds_tar_url 必填")

    body: dict[str, Any] = {
        "data": {
            "dsName": ds_name,
            # 平台接受字符串（"1" / "4"），int 会被拒
            "dsType": str(ds_type),
            "dsSubType": str(ds_sub_type),
            "dsTarUrl": ds_tar_url,
        },
    }
    if name is not None:
        body["data"]["name"] = name
    if ds_ext_info is not None:
        body["data"]["dsExtInfo"] = ds_ext_info

    return api._request("POST", DataHubDsInfoAdd, body=body, wrap=False)


def delete_ds_info(api: AlgAPI, ids: list[int]) -> dict[str, Any]:
    """批量删除数据源（DELETE /ibd-data-hub-web-v2.2/api/ds-info/batchDelete）。

    参数:
      ids: 数据源 ID 列表

    返回: _request 响应（成功时 content 为 {} 或对应消息）。

    注意:
      - 实测发现可能需要先禁用数据源或清空其下位号才能删除
    """
    if not ids:
        return {}
    body = {"data": {"ids": ids}}
    return api._request("DELETE", DataHubDsInfoBatchDelete, body=body, wrap=False)


def change_ds_state(api: AlgAPI, ds_id: int, enabled: bool) -> dict[str, Any]:
    """启用/禁用数据源（POST /ibd-data-hub-web-v2.2/api/ds-info/changeState）。

    参数:
      ds_id:   数据源 ID
      enabled: True=启用(1), False=禁用(0)

    返回: _request 响应 dict。
    """
    body = {"data": {str(ds_id): 1 if enabled else 0}}
    return api._request("POST", DataHubDsInfoChangeState, body=body, wrap=False)


# === 位号分组（tag-group） ===

def add_tag_group(
    api: AlgAPI,
    group_name: str,
    parent_id: str = "0",
) -> dict[str, Any]:
    """创建位号分组节点（POST /ibd-data-hub-web-v2.2/api/tag-group/add）。

    参数:
      group_name: 分组名称
      parent_id:  父分组 ID，默认 "0" = Root

    返回: 新建分组记录，含 id/groupName/parentId/displayIndex/createBy/updateBy/createTime/updateTime
    """
    body = {"data": {"parentId": parent_id, "groupName": group_name}}
    return api._request("POST", DataHubTagGroupAdd, body=body, wrap=False)


def update_tag_group(
    api: AlgAPI,
    group_id: str,
    group_name: str,
    parent_id: str = "0",
) -> dict[str, Any]:
    """编辑位号分组节点（PUT /ibd-data-hub-web-v2.2/api/tag-group/update）。

    参数:
      group_id:   分组 ID
      group_name: 新分组名称
      parent_id:  父分组 ID（可改父节点实现移动），默认 "0" = Root

    返回: 更新后的分组记录（注意：实测返回的 groupName 可能仍是旧值，需 list 确认）
    """
    body = {"data": {"id": group_id, "parentId": parent_id, "groupName": group_name}}
    return api._request("PUT", DataHubTagGroupUpdate, body=body, wrap=False)


def delete_tag_group(
    api: AlgAPI,
    group_ids: list[str],
    is_force: bool = False,
) -> dict[str, Any]:
    """删除位号分组节点（DELETE /ibd-data-hub-web-v2.2/api/tag-group/batchDelete）。

    参数:
      group_ids: 分组 ID 列表（支持批量删除）
      is_force:  True = 同时删除节点下所有位号；False = 只删节点，位号保留

    返回: {isSuccess: bool, success: bool, code: str, msg: str}
    """
    body = {"data": {"groupIds": group_ids, "isForce": is_force}}
    return api._request("DELETE", DataHubTagGroupBatchDelete, body=body, wrap=False)


def get_tag_group_tree(api: AlgAPI) -> dict[str, Any]:
    """获取位号分组节点树（POST /ibd-data-hub-web-v2.2/api/tag-group/groupTree）。

    参数:
      api: AlgAPI 实例

    返回: list[dict]，根节点(id="0", groupName="Root")，每个节点含:
      - id: 分组 ID（字符串）
      - groupName: 分组名
      - parentId: 父分组 ID
      - tagGroupList: 子分组列表（递归）
      - displayIndex: 显示排序
      - createBy/updateBy/createTime/updateTime

    错误情况下返回 _request 响应 dict。
    """
    return api._request("POST", DataHubTagGroupTree, body={}, wrap=False)


def add_tag_group_relation(
    api: AlgAPI,
    group_id: str,
    tag_ids: list[int],
) -> dict[str, Any]:
    """收藏位号--关联位号到分组（POST /ibd-data-hub-web-v2.2/api/tag-group/batchAddRelation）。

    参数:
      group_id: 分组 ID
      tag_ids:  位号 ID 列表

    返回: bool（true=成功）
    """
    body = {"data": {group_id: tag_ids}}
    return api._request("POST", DataHubTagGroupBatchAddRelation, body=body, wrap=False)


def remove_tag_group_relation(
    api: AlgAPI,
    group_id: str,
    tag_ids: list[int],
) -> dict[str, Any]:
    """取消收藏--移除位号与分组的关联（DELETE /ibd-data-hub-web-v2.2/api/tag-group/batchDelRelation）。

    参数:
      group_id: 分组 ID
      tag_ids:  位号 ID 列表

    返回: bool（实测返回 false 但操作实际生效，以 list_favorite_tags 确认为准）
    """
    body = {"data": {group_id: tag_ids}}
    return api._request("DELETE", DataHubTagGroupBatchDelRelation, body=body, wrap=False)


# === 数据源测试（ds-info/test） ===

# testType 枚举
DsTestEnumerate = 1   # 枚举位号（browse UA server 节点）
DsTestReadRT = 2      # 位号实时值（读源端）
DsTestReadRTDB = 3    # 位号实时值（读库）
DsTestHistory = 4     # 历史值（需 beginTime/endTime）
DsTestWrite = 5       # 写值（需 tagValue）


def test_ds_info(
    api: AlgAPI,
    ds_id: int,
    ds_name: str = "",
    test_type: int = DsTestEnumerate,
    tag_name: str | None = None,
    tag_value: str | None = None,
    begin_time: str | None = None,
    end_time: str | None = None,
    interval: int = 0,
    ds_ext_info: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """数据源测试（POST /ibd-data-hub-web-v2.2/api/ds-info/test）。

    在不注册位号的情况下，直接对数据源执行枚举/读/写/历史测试。

    参数:
      ds_id:       数据源 ID
      ds_name:     数据源名（可空）
      test_type:   测试类型（1=枚举 2=读RT 3=读RT库 4=历史 5=写值）
      tag_name:    位号名（testType 2-5 需要）
      tag_value:   写入值（testType=5 需要，字符串形式）
      begin_time:  开始时间（testType=4 需要，格式 "yyyy-MM-dd HH:mm:ss"）
      end_time:    结束时间（testType=4 需要）
      interval:    采样间隔（testType=4 可选）
      ds_ext_info: 数据源扩展信息（默认 {}）

    返回:
      content dict，结构随 testType 变化：
        - testType=1: {successTagNames, successes[{name,browseName,tagDataType,tagDataTypeName,readOnly,...}], total, pageNum, pageSize, totalPage}
        - testType=2/3: {successTagNames, successes[{name,value,quality,timeStamp,dataSourceId}], isAllSuccess}
        - testType=4: {successTagNames, historyValueMap, failTagNames, failMsg, isAllSuccess}
        - testType=5: {successTagNames, failTagNames, failMsg, isAllSuccess}
    """
    data: dict[str, Any] = {
        "dsId": ds_id,
        "dsName": ds_name,
        "testType": test_type,
        "tagValue": tag_value or "",
        "timeStamp": "",
        "dsExtInfo": ds_ext_info or {},
        "beginTime": begin_time or "",
        "endTime": end_time or "",
        "interval": interval,
    }
    if tag_name is not None:
        data["tagName"] = tag_name
    body = {"data": data}
    return api._request("POST", DataHubDsInfoTest, body=body, wrap=False)
