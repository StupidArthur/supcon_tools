"""xlsx 读写抽象.

所有 openpyxl 操作封在这里, 其他模块只看到 list/dict, 不直接接触 openpyxl.

API:
  read_all_sheets(path) -> dict[sheet_name, list[list]]
      读一个 xlsx 全部 sheet, 每 sheet 返回 list of list (第一行是表头).
  read_sheet(path, sheet_name) -> list[list]
      读单个 sheet.
  write_wide_xlsx(path, *, a1, headers, rows, sheet_name) -> path
      写"宽表导入格式" xlsx:
        A1 = a1
        A2 = (空)
        B1.. = headers
        A3.. = rows (每行第一列是 time, 后面是 values)
"""
from openpyxl import Workbook, load_workbook


def read_sheet(path: str, sheet_name: str) -> list[list]:
    """读单个 sheet 全部行, 返回 list[list] (第一行是表头)."""
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb[sheet_name]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def read_all_sheets(path: str) -> dict[str, list[list]]:
    """读全部 sheet, 返回 {sheet_name: rows}. 每 sheet 第一行是表头."""
    wb = load_workbook(path, read_only=True)
    try:
        result = {}
        for name in wb.sheetnames:
            ws = wb[name]
            result[name] = [list(row) for row in ws.iter_rows(values_only=True)]
        return result
    finally:
        wb.close()


def write_wide_xlsx(path: str, *, a1: str, headers: list,
                    rows: list[list], sheet_name: str = "history") -> str:
    """写"宽表导入格式" xlsx (给 importTagValueHistory 用的).

    布局:
      A1 = a1
      A2 = (留空)
      B1..Bn+1 = headers
      A3.. = rows (每行 [time, val_tag1, val_tag2, ...])
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.cell(row=1, column=1, value=a1)
    for i, h in enumerate(headers):
        ws.cell(row=1, column=2 + i, value=h)
    for r_i, row in enumerate(rows):
        for c_i, v in enumerate(row):
            ws.cell(row=3 + r_i, column=1 + c_i, value=v)
    wb.save(path)
    return path
