"""
Excel 导出器（xlsx / xls）

表头与时间列规则与 CSVExporter 一致；默认 sheet 名由调用方传入（缺省「控制器」）。
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from components.utils.logger import get_logger

from .csv_exporter import METADATA_FIELDS
from .template_manager import ExportTemplate, DEFAULT_TIME_DESCRIPTION, DEFAULT_PARAM_DESCRIPTION

logger = get_logger()

# Excel 工作表名称最大长度
EXCEL_SHEET_NAME_MAX_LEN = 31


class ExcelExporter:
    """将快照导出为 .xlsx（openpyxl）或 .xls（xlwt）。"""

    def __init__(
        self,
        template: ExportTemplate,
        file_format: str,
        sheet_name: str = "控制器",
        sample_interval: Optional[float] = None,
    ) -> None:
        fmt = (file_format or "xlsx").lower()
        if fmt not in ("xlsx", "xls"):
            raise ValueError(f"ExcelExporter 仅支持 xlsx/xls，got {file_format}")
        self.template = template
        self.file_format = fmt
        self.sheet_name = (sheet_name or "控制器")[:EXCEL_SHEET_NAME_MAX_LEN]
        self.sample_interval = sample_interval or 1.0

        logger.info(
            "ExcelExporter initialized: format=%s, sheet=%s, header_rows=%d",
            self.file_format,
            self.sheet_name,
            template.header_rows,
        )

    def export(
        self,
        snapshots: List[Dict[str, Any]],
        output_path: str | Path,
        start_time: Optional[float] = None,
        column_keys: Optional[List[str]] = None,
    ) -> None:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        filtered = self._filter_snapshots(snapshots)
        if not filtered:
            logger.warning("没有数据需要导出（Excel）")
            return

        columns = self._determine_columns(filtered[0], column_keys=column_keys)

        if self.file_format == "xlsx":
            self._export_xlsx(output_path, filtered, columns)
        else:
            self._export_xls(output_path, filtered, columns)

        logger.info(
            "Excel 导出成功: 文件=%s, 行数=%d, 列数=%d",
            output_path,
            len(filtered) + self.template.header_rows,
            len(columns) + 1,
        )

    def _filter_snapshots(self, snapshots: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        return [s for s in snapshots if s.get("need_sample", False)]

    def _determine_columns(
        self,
        sample_snapshot: Dict[str, Any],
        column_keys: Optional[List[str]] = None,
    ) -> List[str]:
        if column_keys:
            return [k for k in column_keys if k in sample_snapshot and k not in METADATA_FIELDS]
        return [k for k in sample_snapshot.keys() if k not in METADATA_FIELDS]

    def _display_columns(self, columns: List[str]) -> List[str]:
        # 默认关闭“导出列名转大写”能力，当前统一按原始位号名导出。
        # 如需恢复旧行为，可改回基于 self.template.uppercase_column_names 的分支逻辑。
        return columns

    def _format_time(self, timestamp: float) -> str:
        dt = datetime.fromtimestamp(timestamp)
        return dt.strftime(self.template.time_format)

    def _export_xlsx(
        self,
        output_path: Path,
        snapshots: List[Dict[str, Any]],
        columns: List[str],
    ) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        display_cols = self._display_columns(columns)
        row = 1

        # 第一行表头
        ws.cell(row=row, column=1, value=self.template.time_column_name)
        for i, name in enumerate(display_cols, start=2):
            ws.cell(row=row, column=i, value=name)
        row += 1

        if self.template.header_rows == 2:
            time_desc = self.template.time_row2_description or DEFAULT_TIME_DESCRIPTION
            ws.cell(row=row, column=1, value=time_desc)
            for i in range(len(columns)):
                ws.cell(row=row, column=i + 2, value=DEFAULT_PARAM_DESCRIPTION)
            row += 1

        for snap in snapshots:
            time_str = self._format_time(float(snap.get("sim_time", 0.0)))
            ws.cell(row=row, column=1, value=time_str)
            for i, col in enumerate(columns, start=2):
                val = snap.get(col, "")
                if val is None:
                    ws.cell(row=row, column=i, value="")
                else:
                    ws.cell(row=row, column=i, value=val)
            row += 1

        wb.save(output_path)

    def _export_xls(
        self,
        output_path: Path,
        snapshots: List[Dict[str, Any]],
        columns: List[str],
    ) -> None:
        import xlwt

        book = xlwt.Workbook(encoding="utf-8")
        sheet = book.add_sheet(self.sheet_name)

        display_cols = self._display_columns(columns)
        row = 0

        sheet.write(row, 0, self.template.time_column_name)
        for i, name in enumerate(display_cols):
            sheet.write(row, i + 1, name)
        row += 1

        if self.template.header_rows == 2:
            time_desc = self.template.time_row2_description or DEFAULT_TIME_DESCRIPTION
            sheet.write(row, 0, time_desc)
            for i in range(len(columns)):
                sheet.write(row, i + 1, DEFAULT_PARAM_DESCRIPTION)
            row += 1

        for snap in snapshots:
            time_str = self._format_time(float(snap.get("sim_time", 0.0)))
            sheet.write(row, 0, time_str)
            for i, col in enumerate(columns):
                val = snap.get(col, "")
                if val is None:
                    sheet.write(row, i + 1, "")
                else:
                    sheet.write(row, i + 1, str(val))
            row += 1

        book.save(str(output_path))
