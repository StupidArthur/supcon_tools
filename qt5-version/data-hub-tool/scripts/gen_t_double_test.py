import _path_helper  # noqa: 帮脚本 import 根目录的模块

#!/usr/bin/env python3
"""
为 t_double_N..M (N 个空 tag) 生成 long 格式测试 xlsx.
格式跟 export_data.xlsx 一致 (字符串 Tag Value), 给 migrate.py 读.

用法:
  python gen_t_double_test.py                                       # 默认 t_double_13..20, 10000 点, 2026-10-15
  python gen_t_double_test.py 21 10                                # t_double_21..30, 10000 点, 2026-10-15
  python gen_t_double_test.py 21 10 1000 2026-11-01               # t_double_21..30, 1000 点, 2026-11-01
  python gen_t_double_test.py 31 20 500 2026-12-01 my_test.xlsx   # 自定义输出文件名
"""
import os
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook

DEFAULT_TAG_START = 13
DEFAULT_N_TAGS = 8
DEFAULT_N_POINTS = 10000
DEFAULT_START = "2026-10-15"
DEFAULT_PERIOD_S = 10
DEFAULT_OUT = "t_double_test_export.xlsx"


def gen(tag_start: int, n_tags: int, n_points: int, start_str: str,
        period_s: int, out_xlsx: str):
    start = datetime.strptime(start_str, "%Y-%m-%d")
    end = start + timedelta(seconds=period_s * (n_points - 1))
    wb = Workbook()
    wb.remove(wb.active)
    for i in range(tag_start, tag_start + n_tags):
        tag = f"t_double_{i}"
        ws = wb.create_sheet(title=tag)
        # 表头
        ws.cell(1, 1, "Tag Time")
        ws.cell(1, 2, "App Time")
        ws.cell(1, 3, "Quality")
        ws.cell(1, 4, "Tag Value")
        # 数据: 倒序 (新→旧)
        for r in range(n_points):
            idx = n_points - 1 - r
            t = start + timedelta(seconds=period_s * idx)
            val = idx % 100  # 0~99 循环
            ws.cell(r + 2, 1, t.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(r + 2, 2, t.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(r + 2, 3, 192)
            ws.cell(r + 2, 4, f"{val}.0")  # 字符串 + .0, 匹配真实导出格式
    wb.save(out_xlsx)
    print(f"[gen] 已生成 {out_xlsx}")
    print(f"  {n_tags} 个 tag: t_double_{tag_start}..t_double_{tag_start + n_tags - 1}")
    print(f"  每个 {n_points} 点, 周期 {period_s}s, 时间 {start} ~ {end}")
    print(f"  值: 0~99 循环 (字符串), 方向: 倒序 (新→旧)")


if __name__ == "__main__":
    tag_start = int(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_TAG_START
    n_tags = int(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_N_TAGS
    n_points = int(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_N_POINTS
    start_str = sys.argv[4] if len(sys.argv) > 4 else DEFAULT_START
    out_xlsx = sys.argv[5] if len(sys.argv) > 5 else DEFAULT_OUT
    gen(tag_start, n_tags, n_points, start_str, DEFAULT_PERIOD_S, out_xlsx)
